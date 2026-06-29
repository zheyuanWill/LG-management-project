from __future__ import annotations

import re
from copy import copy
from dataclasses import dataclass
from io import BytesIO
from typing import Dict, Iterable, List, Optional, Tuple

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell


@dataclass(frozen=True)
class QuoteExcelConfig:
    steel_sheet_keywords: Tuple[str, ...] = ("steel",)
    steel_part_keyword: str = "PART1"
    steel_style_row: int = 30
    mach_sheet_keywords: Tuple[str, ...] = ("mach",)
    mach_tariff_sheet_name: str = "Quote"
    mach_tariff_row_range: Tuple[int, int] = (1, 2000)
    preserve_marker: str = "THE LIST IS UNTIL"


def _norm_text(v) -> str:
    if v is None:
        return ""
    return str(v).strip()


def _find_sheet_by_keywords(wb, keywords: Tuple[str, ...]) -> str:
    lowered_keywords = tuple(k.lower() for k in keywords if k and str(k).strip())
    if not lowered_keywords:
        return wb.sheetnames[0]
    for name in wb.sheetnames:
        lowered = name.lower()
        if all(k in lowered for k in lowered_keywords):
            return name
    return wb.sheetnames[0]


def _find_header_row(ws, required_keywords: Iterable[str], scan_limit: int = 300) -> Optional[int]:
    required = [k.upper() for k in required_keywords]
    row_idx = 0
    for row in ws.iter_rows(values_only=True):
        row_idx += 1
        if row_idx > scan_limit:
            break
        joined = "|".join(_norm_text(v) for v in (row[:20] if row else [])).upper()
        if all(k in joined for k in required):
            return row_idx
    return None


def _find_preserve_row(ws, marker: str) -> Optional[int]:
    marker_u = (marker or "").upper()
    if not marker_u:
        return None
    row_idx = 0
    for row in ws.iter_rows(values_only=True):
        row_idx += 1
        joined = "|".join(_norm_text(v) for v in (row[:30] if row else [])).upper()
        if marker_u in joined:
            return row_idx
    return None


def _set_cell_value(ws, row: int, col: int, value):
    cell = ws.cell(row, col)
    if isinstance(cell, MergedCell):
        return
    cell.value = value


def _clear_range(ws, start_row: int, end_row: int, cols: Iterable[int]):
    for r in range(start_row, end_row + 1):
        for c in cols:
            _set_cell_value(ws, r, c, None)


def _safe_write_with_style(ws, row: int, col: int, value, style_cell):
    cell = ws.cell(row, col)
    if isinstance(cell, MergedCell):
        return False
    cell.value = value
    if style_cell and style_cell.has_style:
        cell.font = copy(style_cell.font)
        cell.border = copy(style_cell.border)
        cell.fill = copy(style_cell.fill)
        cell.number_format = copy(style_cell.number_format)
        cell.protection = copy(style_cell.protection)
        cell.alignment = copy(style_cell.alignment)
    return True


def _find_source_start_steel(ws) -> int:
    for row in ws.iter_rows():
        for cell in row:
            if cell.value:
                txt = str(cell.value)
                if "Description" in txt or "DescripTion" in txt:
                    return cell.row + 1
    return 11


def _find_template_start_steel(ws) -> int:
    for row in ws.iter_rows():
        values = []
        for c in row:
            if c.value:
                values.append(str(c.value).strip().upper())
            else:
                values.append("")
        text = "|".join(values)
        if "REF" in text and ("STEEL" in text or "JOB" in text or "DRY" in text):
            return row[0].row + 1
    return 30


def _find_preserve_row_steel(ws, marker: str) -> Optional[int]:
    marker_u = (marker or "").upper()
    if not marker_u:
        return None
    for row in range(1, ws.max_row + 1):
        row_values = []
        for col in range(1, ws.max_column + 1):
            value = ws.cell(row, col).value
            if value is not None and str(value).strip():
                row_values.append((col, str(value).strip()))
        if not row_values:
            continue
        if any(marker_u in text.upper() for _, text in row_values):
            return row
    return None


def _clear_data_in_range_steel(ws, start_row: int, end_row: int):
    for row in range(start_row, end_row + 1):
        for col in range(2, 6):
            cell = ws.cell(row, col)
            if not isinstance(cell, MergedCell):
                cell.value = None


def _normalize_uom(uom):
    if uom is None:
        return None
    return str(uom).strip().lower()


def _parse_t_material(text: str):
    sets_match = re.search(r"(\d+)\s*sets?", text, re.I)
    sets_multiplier = int(sets_match.group(1)) if sets_match else 1

    m = re.search(
        r"T\s*(\d+(?:\.\d+)?)\s*[*xX×]\s*(?:Ø\s*)?(\d+(?:\.\d+)?)\s*[*xX×]\s*(?:Ø\s*)?(\d+(?:\.\d+)?)\s*[*xX×]\s*(\d+)\s*pcs",
        text,
        re.I,
    )
    if m:
        thickness = float(m.group(1))
        width = float(m.group(2)) / 1000
        length = float(m.group(3)) / 1000
        pcs = int(m.group(4))
        qty = thickness * width * length * 8 * pcs * sets_multiplier
        return round(qty, 3), "kg"

    compact = re.search(r"T\s*(\d+)\s*pcs", text, re.I)
    if compact:
        digits = compact.group(1)
        parsed = None
        for thickness_len in (1, 2):
            for pcs_len in (1, 2):
                width_len = 3
                length_len = len(digits) - thickness_len - width_len - pcs_len
                if length_len not in (3, 4):
                    continue
                thickness = int(digits[:thickness_len])
                width = int(digits[thickness_len:thickness_len + width_len])
                length = int(digits[thickness_len + width_len:thickness_len + width_len + length_len])
                pcs = int(digits[-pcs_len:])
                if width < 100:
                    continue
                parsed = (thickness, width, length, pcs)
                break
            if parsed:
                break
        if parsed:
            thickness, width, length, pcs = parsed
            qty = thickness * (width / 1000) * (length / 1000) * 8 * pcs * sets_multiplier
            return round(qty, 3), "kg"
    return None


def _parse_l_material(text: str):
    m = re.search(r"^L.*?(\d+)\D*?(\d+)(?:pcs|sets)", text, re.I)
    if m:
        length = int(m.group(1))
        pcs = int(m.group(2))
        if length > 1000:
            qty = (length / 1000) * pcs
        else:
            qty = 1 * pcs
        return round(qty, 3), "M"
    return None


def _parse_hp_material(text: str):
    m = re.search(r"^HP.*?(\d+)\D*?(\d+)(?:pcs|sets)", text, re.I)
    if m:
        length = int(m.group(1))
        pcs = int(m.group(2))
        if length > 1000:
            qty = (length / 1000) * pcs
        else:
            qty = 1 * pcs
        return round(qty, 3), "PCS"
    return None


def _parse_pipe_material(text: str):
    m = re.search(r"(?:Ø\s*\d+(?:\.\d+)?\s*[*xX×]\s*)?(\d+(?:\.\d+)?)\s*[*xX×]\s*(\d+)\s*pcs", text, re.I)
    if m:
        length = float(m.group(1))
        pcs = int(m.group(2))
        if length > 1000:
            qty = (length / 1000) * pcs
            uom = "m"
        else:
            qty = pcs
            uom = "pc"
        return round(qty, 3), uom

    compact = re.search(r"Ø?\s*(\d+)\s*pcs", text, re.I)
    if compact:
        digits = compact.group(1)
        parsed = None
        for spec_len in (2, 3, 4):
            for pcs_len in (1, 2):
                length_len = len(digits) - spec_len - pcs_len
                if length_len not in (3, 4):
                    continue
                length = int(digits[spec_len:spec_len + length_len])
                pcs = int(digits[-pcs_len:])
                parsed = (length, pcs)
                break
            if parsed:
                break
        if parsed:
            length, pcs = parsed
            if length > 1000:
                qty = (length / 1000) * pcs
                uom = "m"
            else:
                qty = pcs
                uom = "pc"
            return round(qty, 3), uom
    return None


def _parse_tonnage_as_pcs(text: str):
    m = re.search(r"(\d+(?:\.\d+)?)\s*T\s*[*xX×]\s*(\d+)\s*pcs", text, re.I)
    if m:
        return int(m.group(2)), "pcs"
    return None


def _parse_hours_job(text: str):
    m = re.search(r"(\d+)\s*pcs\s*[*xX×]\s*(\d+(?:\.\d+)?)\s*hours\s*[*xX×]\s*(\d+)\s*tank\s*[*xX×]\s*(\d+)\s*sets", text, re.I)
    if m:
        return float(m.group(2)), "hours"
    return None


def _parse_safety_watch(text: str):
    m = re.search(r"(\d+)\s*(?:person|pcs?)\s*[*xX×]?\s*(\d+)\s*hours.*?(\d+)\s*days", text, re.I)
    if m:
        hours = int(m.group(2))
        days = int(m.group(3))
        if hours <= 8:
            factor = 1
        elif hours < 16:
            factor = 1.5
        else:
            factor = 2
        return factor * days, "DAYS"
    m = re.search(r"(\d+)\s*(?:person|pcs?)\s*[*xX×]?\s*(\d+)\s*hours", text, re.I)
    if m:
        hours = int(m.group(2))
        if hours <= 8:
            factor = 1
        elif hours < 16:
            factor = 1.5
        else:
            factor = 2
        return factor, "DAYS"
    return None


def _parse_fireproof(text: str):
    m = re.search(r"(\d+(?:\.\d+)?)\s*m2", text, re.I)
    if m:
        return float(m.group(1)), "M2"
    return None


def _parse_ut(text: str):
    m = re.search(r"UT\s*grinding[:：]?\s*(\d+(?:\.\d+)?)\s*[*xX×]\s*(\d+(?:\.\d+)?)\s*[*xX×]\s*(\d+)\s*pcs", text, re.I)
    if m:
        width = float(m.group(1)) / 1000
        length = float(m.group(2)) / 1000
        pcs = int(m.group(3))
        return round((width + length) * 2 * pcs, 3), "m"
    m = re.search(r"UT\s*grinding[:：]?\s*(\d+(?:\.\d+)?)\s*[*xX×]\s*(\d+(?:\.\d+)?)", text, re.I)
    if m:
        width = float(m.group(1)) / 1000
        length = float(m.group(2)) / 1000
        return round((width + length) * 2, 3), "m"
    m = re.search(r"UT[:：]?\s*(\d+)", text, re.I)
    if m:
        value = int(m.group(1))
        if value > 1000:
            return round(value / 1000, 3), "m"
        return 1, "pc"
    return None


def _parse_vt(text: str):
    m = re.search(r"VT[:：]?\s*(\d+)", text, re.I)
    if m:
        value = int(m.group(1))
        if value > 1000:
            return round(value / 1000, 3), "m"
        return 1, "pc"
    return None


def _parse_mt(text: str):
    m = re.search(r"MT[:：]?\s*(\d+)", text, re.I)
    if m:
        value = int(m.group(1))
        if value > 1000:
            return round(value / 1000, 3), "m"
        return 1, "pc"
    return None


def _parse_pt(text: str):
    m = re.search(r"PT[:：]?\s*(\d+(?:\.\d+)?)", text, re.I)
    if m:
        try:
            value = float(m.group(1))
            if value > 1000:
                return round(value / 1000, 3), "m"
            return 1, "pc"
        except Exception:
            pass
    m = re.search(r"PT[:：]?\s*(\d+)", text, re.I)
    if m:
        value = int(m.group(1))
        if value > 1000:
            return round(value / 1000, 3), "m"
        return 1, "pc"
    return None


def _parse_lifting_lug(text: str):
    m = re.search(r"lifting\s+lug.*?(\d+)\s*pcs", text, re.I)
    if m:
        return int(m.group(1)), "pc"
    return None


def _parse_tonnage(text: str):
    m = re.search(r"(\d+(?:\.\d+)?)\s*T\D*?(\d+)(?:pcs|sets)", text, re.I)
    if m:
        ton = float(m.group(1))
        pcs = int(m.group(2))
        return ton * pcs, "t"
    m = re.search(r"(\d+(?:\.\d+)?)\s*T", text, re.I)
    if m:
        return float(m.group(1)), "t"
    return None


def _parse_fallback(text: str):
    m = re.search(r"(\d+(?:\.\d+)?)\s*([A-Za-z0-9]+)\s*\*?\s*(\d+)", text, re.I)
    if m:
        base = float(m.group(1))
        uom = m.group(2)
        multiplier = int(m.group(3))
        if uom.lower() in ["set", "sets"]:
            uom = "pc"
        return base * multiplier, uom
    m = re.search(r"(\d+(?:\.\d+)?)\s*([A-Za-z0-9]+)", text, re.I)
    if m:
        return float(m.group(1)), m.group(2)
    return None


def _parse_item(text: str):
    txt = str(text).strip()
    funcs = [
        _parse_hours_job,
        _parse_safety_watch,
        _parse_fireproof,
        _parse_pt,
        _parse_ut,
        _parse_vt,
        _parse_mt,
        _parse_tonnage_as_pcs,
        _parse_lifting_lug,
        _parse_tonnage,
        _parse_t_material,
        _parse_l_material,
        _parse_hp_material,
        _parse_pipe_material,
        _parse_fallback,
    ]
    for f in funcs:
        result = f(txt)
        if result:
            qty, uom = result
            return qty, _normalize_uom(uom)
    return None, None


def _import_steel(template_wb, steel_source_bytes: bytes, config: QuoteExcelConfig):
    src_wb = load_workbook(BytesIO(steel_source_bytes), data_only=True)
    src_ws = None
    for sheet_name in src_wb.sheetnames:
        if config.steel_part_keyword and config.steel_part_keyword in sheet_name:
            src_ws = src_wb[sheet_name]
            break
    if not src_ws:
        src_ws = src_wb.active

    source_start = _find_source_start_steel(src_ws)

    all_data = []
    ref_no = 0
    previous_cell_empty = True
    for row in range(source_start, src_ws.max_row + 1):
        cell_value = src_ws.cell(row, 3).value
        current_cell_empty = not cell_value or not str(cell_value).strip()
        if current_cell_empty:
            previous_cell_empty = True
            continue
        text = str(cell_value).strip()
        if previous_cell_empty:
            ref_no += 1
            all_data.append((ref_no, text, None, None))
        else:
            qty, uom = _parse_item(text)
            all_data.append((None, text, qty, uom))
        previous_cell_empty = False

    steel_sheet_name = _find_sheet_by_keywords(template_wb, config.steel_sheet_keywords)
    ws = template_wb[steel_sheet_name]

    target_start = _find_template_start_steel(ws)
    preserve_row = _find_preserve_row_steel(ws, config.preserve_marker)
    if not preserve_row:
        raise RuntimeError("Could not find preserve marker row in steel sheet.")

    available_rows = preserve_row - target_start
    if len(all_data) > available_rows:
        insert_count = len(all_data) - available_rows
        ws.insert_rows(preserve_row, insert_count)
        preserve_row += insert_count

    _clear_data_in_range_steel(ws, target_start, preserve_row - 1)

    style_row = max(int(config.steel_style_row or 1), 1)
    style_cell_ref = ws.cell(style_row, 2)
    style_cell_desc = ws.cell(style_row, 3)
    style_cell_qty = ws.cell(style_row, 4)
    style_cell_uom = ws.cell(style_row, 5)

    current_row = target_start
    for ref_no_val, desc, qty, uom in all_data:
        if current_row > ws.max_row:
            ws.insert_rows(current_row)
        if ref_no_val is not None:
            _safe_write_with_style(ws, current_row, 2, ref_no_val, style_cell_ref)
        _safe_write_with_style(ws, current_row, 3, desc, style_cell_desc)
        if qty is not None:
            _safe_write_with_style(ws, current_row, 4, qty, style_cell_qty)
        if uom is not None:
            _safe_write_with_style(ws, current_row, 5, uom, style_cell_uom)
        current_row += 1

    delete_count = preserve_row - current_row
    if delete_count > 0:
        ws.delete_rows(current_row, delete_count)


@dataclass(frozen=True)
class _TariffRow:
    straight_pipe: float
    bend_add: float
    flange: float


def _parse_tariff_table(tariff_bytes: bytes, config: QuoteExcelConfig) -> Dict[str, _TariffRow]:
    wb = load_workbook(BytesIO(tariff_bytes), read_only=True, data_only=True)
    if config.mach_tariff_sheet_name and config.mach_tariff_sheet_name in wb.sheetnames:
        ws = wb[config.mach_tariff_sheet_name]
    else:
        ws = wb[wb.sheetnames[0]]

    start, end = config.mach_tariff_row_range
    start = max(int(start or 1), 1)
    end = max(int(end or start), start)

    table: Dict[str, _TariffRow] = {}
    for r in range(start, end + 1):
        row_text = _norm_text(ws.cell(r, 3).value)
        if not row_text:
            continue
        m = re.search(r"^\s*(<\s*25|\d+)\s+(\d+(?:\.\d+)?)\s+(\d+(?:\.\d+)?)\s+(\d+(?:\.\d+)?)\s*$", row_text)
        if not m:
            continue
        diameter_key = m.group(1).replace(" ", "")
        straight = float(m.group(2))
        bend = float(m.group(3))
        flange = float(m.group(4))
        table[diameter_key] = _TariffRow(straight_pipe=straight, bend_add=bend, flange=flange)
    if not table:
        raise RuntimeError("Could not parse tariff table.")
    return table


def _lookup_tariff(table: Dict[str, _TariffRow], diameter: int) -> Optional[_TariffRow]:
    if diameter <= 25 and "<25" in table:
        return table["<25"]
    return table.get(str(diameter))


def _shift_diameter_up(table: Dict[str, _TariffRow], diameter: int, steps: int) -> int:
    ordered = [25, 40, 50, 65, 80, 90, 100, 125, 150, 200, 250, 300, 350, 400, 450]
    current = 25 if diameter < 25 else diameter
    if current not in ordered:
        for value in ordered:
            if value >= current:
                current = value
                break
        else:
            current = ordered[-1]
    idx = ordered.index(current)
    idx = min(idx + max(steps, 0), len(ordered) - 1)
    shifted = ordered[idx]
    if str(shifted) in table:
        return shifted
    return diameter


def _try_calc_unit_usd(text: str, table: Dict[str, _TariffRow]) -> Optional[float]:
    cleaned = text.replace("×", "*").replace("X", "*").replace("x", "*")
    m = re.search(r"(\d+)\s*A", cleaned, re.I)
    if not m:
        return None
    diameter = int(m.group(1))
    s_shift = 0
    for count_text in re.findall(r"(\d+)\s*S\b", cleaned, re.I):
        s_shift += int(count_text)
    if s_shift:
        diameter = _shift_diameter_up(table, diameter, s_shift)
    tariff = _lookup_tariff(table, diameter)
    if not tariff:
        return None

    tail = cleaned[m.end():]
    parts = [p.strip() for p in tail.split("*") if p.strip()]

    length_mm: Optional[float] = None
    counts: List[Tuple[str, int]] = []
    for p in parts:
        m_count = re.fullmatch(r"(\d+)\s*([FBES])", p, re.I)
        if m_count:
            counts.append((m_count.group(2).upper(), int(m_count.group(1))))
            continue
        if length_mm is None:
            m_len = re.fullmatch(r"\d+(?:\.\d+)?", p)
            if m_len:
                length_mm = float(p)

    value = tariff.straight_pipe if length_mm is None else tariff.straight_pipe * (length_mm / 1000.0)
    for code, count in counts:
        if code == "F":
            value += count * tariff.flange
        elif code == "E":
            value += count * tariff.bend_add
        elif code == "B":
            value += count * tariff.straight_pipe

    return round(value, 3)


def _import_mach(template_wb, mach_wdr_bytes: bytes, mach_tariff_bytes: bytes, config: QuoteExcelConfig):
    tariff_table = _parse_tariff_table(mach_tariff_bytes, config)

    source_wb = load_workbook(BytesIO(mach_wdr_bytes), read_only=True, data_only=True)
    source_ws = source_wb[source_wb.sheetnames[0]]
    source_header_row = _find_header_row(source_ws, required_keywords=["S NO", "DESCRIPTION", "QTY", "UOM"], scan_limit=2000)
    if not source_header_row:
        raise RuntimeError("Could not find header row in WDR source sheet.")
    source_start_row = source_header_row + 1

    mach_sheet_name = _find_sheet_by_keywords(template_wb, config.mach_sheet_keywords)
    ws = template_wb[mach_sheet_name]

    ref_header_row = None
    for r in range(1, min(ws.max_row, 500) + 1):
        row_vals = [_norm_text(ws.cell(r, c).value) for c in range(1, 10)]
        joined = "|".join(row_vals).upper()
        if "REF NO" in joined:
            ref_header_row = r
            break
    if not ref_header_row:
        raise RuntimeError("Could not find 'Ref No' header row in target mach sheet.")

    data_start_row = ref_header_row + 1
    preserve_row = _find_preserve_row(ws, config.preserve_marker)
    if not preserve_row:
        raise RuntimeError("Could not find preserve marker row in mach sheet.")

    REF_COL = 2
    DESC_COL = 3
    QTY_COL = 4
    UOM_COL = 5
    UNIT_COL = 6

    if _norm_text(ws.cell(data_start_row, REF_COL).value).upper() == "M":
        data_start_row += 1

    _clear_range(ws, data_start_row, preserve_row - 1, cols=[DESC_COL, QTY_COL, UOM_COL, UNIT_COL])

    max_r = preserve_row - 1
    target_capacity = max_r - data_start_row + 1

    source_iter = source_ws.iter_rows(
        min_row=source_start_row,
        max_row=source_start_row + target_capacity - 1,
        max_col=5,
        values_only=True,
    )
    for offset, row in enumerate(source_iter):
        target_r = data_start_row + offset
        if target_r > max_r:
            break

        desc = _norm_text(row[2] if len(row) > 2 else None)
        qty_val = row[3] if len(row) > 3 else None
        uom = _norm_text(row[4] if len(row) > 4 else None)

        qty: Optional[float] = None
        if qty_val is not None and _norm_text(qty_val) != "":
            try:
                qty = float(qty_val)
            except Exception:
                qty = None

        _set_cell_value(ws, target_r, DESC_COL, desc if desc else None)
        _set_cell_value(ws, target_r, QTY_COL, qty)
        _set_cell_value(ws, target_r, UOM_COL, uom if uom else None)

        if desc:
            unit_value = _try_calc_unit_usd(desc, tariff_table)
            if unit_value is not None:
                _set_cell_value(ws, target_r, UNIT_COL, unit_value)


def generate_quote_excel_bytes(
    *,
    template_bytes: bytes,
    steel_source_bytes: bytes,
    mach_wdr_bytes: bytes,
    mach_tariff_bytes: bytes,
    config: Optional[QuoteExcelConfig] = None,
) -> bytes:
    if not template_bytes:
        raise RuntimeError("Template excel is empty.")
    if not steel_source_bytes:
        raise RuntimeError("Steel source excel is empty.")
    if not mach_wdr_bytes:
        raise RuntimeError("Mach WDR excel is empty.")
    if not mach_tariff_bytes:
        raise RuntimeError("Mach tariff excel is empty.")

    cfg = config or QuoteExcelConfig()
    wb = load_workbook(BytesIO(template_bytes))
    _import_steel(wb, steel_source_bytes, cfg)
    _import_mach(wb, mach_wdr_bytes, mach_tariff_bytes, cfg)
    out = BytesIO()
    wb.save(out)
    return out.getvalue()

